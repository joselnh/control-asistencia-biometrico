import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
import sys
import datetime
import re

def vista_previa_y_guardar(df):
    # Usamos tabulaciones para la vista previa básica, ya que no tiene nombre de columnas en sí porque están en la fila 0
    text_content = df.to_string(index=False, header=False)

    # Crear ventana de vista previa usando Toplevel
    preview_win = tk.Toplevel(root)
    preview_win.title("Vista previa del reporte consolidado (Excel)")
    preview_win.geometry("900x650")
    
    lbl = tk.Label(preview_win, text="Vista previa de los datos a exportar a Excel:", font=("Arial", 12, "bold"))
    lbl.pack(pady=10)
    
    frame = tk.Frame(preview_win)
    frame.pack(expand=True, fill="both", padx=10, pady=5)
    
    scrollbar_y = tk.Scrollbar(frame)
    scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
    
    scrollbar_x = tk.Scrollbar(frame, orient=tk.HORIZONTAL)
    scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
    
    text_area = tk.Text(frame, wrap=tk.NONE, yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set, font=("Courier New", 10))
    text_area.insert(tk.END, text_content)
    text_area.config(state=tk.DISABLED) 
    text_area.pack(expand=True, fill="both")
    
    scrollbar_y.config(command=text_area.yview)
    scrollbar_x.config(command=text_area.xview)

    # Botón para guardar
    btn_guardar = tk.Button(
        preview_win, 
        text="Guardar en Excel", 
        command=lambda: guardar_excel(df, preview_win), 
        font=("Arial", 11, "bold"), 
        bg="#4CAF50", 
        fg="white"
    )
    btn_guardar.pack(pady=10)

    # Cuando se cierre la ventana de vista previa, destruir root para salir del programa
    preview_win.protocol("WM_DELETE_WINDOW", root.destroy)

def guardar_excel(df, parent_win):
    # Pide la ruta de guardado
    file_path = filedialog.asksaveasfilename(
        parent=parent_win,
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")]
    )

    if not file_path:
        return

    # ---------------------------------------------------------------
    # ✅ PARTE 1: Crear el archivo Excel limpio (sin colores ni fórmulas extra)
    # ---------------------------------------------------------------
    df.to_excel(file_path, index=False, header=False)

    # ---------------------------------------------------------------
    # ✅ PARTE 2: Abrir Excel con openpyxl para colorear las celdas y formatear
    # ---------------------------------------------------------------
    import openpyxl
    from openpyxl.styles import PatternFill

    # Cargar el libro que acabamos de guardar
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Crear un relleno rojo sólido
    red_fill = PatternFill(
        start_color="FFFF0000",
        end_color="FFFF0000",
        fill_type="solid"
    )

    # Recorrer las filas del Excel para colorear y formatear por empleado
    # ws.iter_rows() usa índices base 1 (la primera fila real es la 1)
    for row_index, row in enumerate(ws.iter_rows(), start=1):
        # El primer día siempre está en la columna 1
        day_value = ws.cell(row=row_index, column=1).value
        is_labor_day = day_value in ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]

        # Recorremos los empleados en la fila
        num_empleados = (ws.max_column - 3) // 6
        for i in range(num_empleados):
            start_col = 4 + i * 6
            
            cell_ingreso = ws.cell(row=row_index, column=start_col + 1)
            cell_salida = ws.cell(row=row_index, column=start_col + 2)
            cell_extras = ws.cell(row=row_index, column=start_col + 3)
            cell_debe = ws.cell(row=row_index, column=start_col + 4)
            
            ing_empty = (cell_ingreso.value is None or str(cell_ingreso.value).strip() == "")
            sal_empty = (cell_salida.value is None or str(cell_salida.value).strip() == "")
            
            # Solo colorear y colocar guiones en días laborables (Lunes a Sábado)
            if is_labor_day:
                # 1. Pintar de rojo si están vacías
                if ing_empty:
                    cell_ingreso.fill = red_fill
                if sal_empty:
                    cell_salida.fill = red_fill
                
                # 2. Rellenar vacías con "-" si no hay vacías en ingreso/salida
                if not ing_empty and not sal_empty:
                    if cell_extras.value is None or str(cell_extras.value).strip() == "":
                        cell_extras.value = "---"
                    if cell_debe.value is None or str(cell_debe.value).strip() == "":
                        cell_debe.value = "---"
            
            # 3. Formatear celdas de tiempo para todos los días (incluidos domingos) si tienen marcas
            for cell_temp, rel_type in [(cell_ingreso, "time"), (cell_salida, "time"), (cell_extras, "duration"), (cell_debe, "duration")]:
                val = cell_temp.value
                if val is not None and str(val).strip() not in ["", "-"]:
                    match = re.match(r"^(\d+):(\d+)(?::(\d+))?$", str(val).strip())
                    if match:
                        h = int(match.group(1))
                        m = int(match.group(2))
                        day_fraction = (h * 60 + m) / 1440.0
                        cell_temp.value = day_fraction
                        
                        if rel_type == "time":
                            cell_temp.number_format = 'hh:mm'
                        else:
                            cell_temp.number_format = '[h]:mm'

    # Formatear la fila de Totales
    totals_row_idx = ws.max_row
    tot_row = ws[totals_row_idx]
    for col_index, cell in enumerate(tot_row, start=1):
        if col_index >= 4:
            rel_col = (col_index - 4) % 6
            if rel_col in [3, 4]:  # H. Extras, H. Debe (los de sumas)
                cell.number_format = '[h]:mm'

    # Activar la protección de la hoja y aplicar bloqueo selectivo (bloquear no nulas y no fórmulas)
    from openpyxl.styles import Protection
    ws.protection.sheet = True
    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            if val is None or str(val).strip() == "" or str(val).startswith("="):
                cell.protection = Protection(locked=False)
            else:
                cell.protection = Protection(locked=True)

    # Guardar el archivo Excel modificado
    wb.save(file_path)

    messagebox.showinfo("Guardado", "¡Excel generado con éxito!", parent=parent_win)

def a_tiempo(s):
    if not s: return ""
    try: return datetime.datetime.strptime(s, "%H:%M").time()
    except Exception: return ""


def format_td(segundos):
    h = int(abs(segundos) // 3600)
    m = int((abs(segundos) % 3600) // 60)
    return datetime.time(h, m)

def calcular_horas(fecha_str, entrada_raw, salida_raw):

    if not entrada_raw or not salida_raw:
        return "", ""
    try:
        fecha_obj = datetime.datetime.strptime(fecha_str, "%Y-%m-%d")
        dia_semana = fecha_obj.weekday()
        t_in = datetime.datetime.strptime(entrada_raw, "%H:%M")
        t_out = datetime.datetime.strptime(salida_raw, "%H:%M")
        
        if dia_semana < 5:
            std_in = hora_entrada_std
            std_out = hora_salida_std
        elif dia_semana == 5:
            std_in = hora_entrada_std
            std_out = hora_salida_sabado
        else:
            std_in = t_in
            std_out = t_in
            
        trabajadas_td = t_out - t_in
        std_td = std_out - std_in
        diff_segundos = trabajadas_td.total_seconds() - std_td.total_seconds()
                  
        if diff_segundos > 0:
            # Si es domingo (dia_semana == 6), se cuenta cualquier marcación.
            # De lunes a sábado (dia_semana < 6), solo se cuenta si es >= 15 minutos (900 segundos).
            if dia_semana == 6 or diff_segundos >= 900:
                return format_td(diff_segundos), ""
            else:
                return "", ""
        elif diff_segundos < 0:
            return "", format_td(diff_segundos)
        else:
            return "", ""
    except Exception:
        return "", ""

# Define horas de entrada y salida estándar
hora_entrada_std=datetime.datetime.strptime("08:00", "%H:%M")
hora_salida_std=datetime.datetime.strptime("17:30", "%H:%M")
hora_salida_sabado=datetime.datetime.strptime("13:00", "%H:%M")

# 1. Cargar el archivo .xls con pandas usando el motor xlrd de manera condicional
root = tk.Tk()
root.withdraw()  # Oculta la ventana principal de tkinter

file_path = filedialog.askopenfilename(
    title="Seleccione el archivo de reporte de asistencia (.xls / .xlsx)",
    filetypes=[("Archivos de Excel", "*.xls *.xlsx"), ("Todos los archivos", "*.*")]
)

if not file_path:
    print("No se seleccionó ningún archivo.")
    sys.exit()

# Leemos el archivo sin cabeceras automáticas para manejar nosotros las filas
if file_path.lower().endswith('.xlsx'):
    df_raw = pd.read_excel(file_path, sheet_name="Reporte de Asistencia", header=None)
else:
    df_raw = pd.read_excel(file_path, sheet_name="Reporte de Asistencia", header=None, engine='xlrd')
# 2. Buscar el período de manera dinámica
anio_mes = "2026-05"  # Valor por defecto
for r_idx in range(min(10, len(df_raw))):
    row_series = df_raw.iloc[r_idx]
    periodo_labels = row_series[row_series.astype(str).str.contains("Periodo:", case=False, na=False)]
    if not periodo_labels.empty:
        col_idx = periodo_labels.index[0]
        # Buscar el siguiente valor no nulo en esa fila
        for c in range(col_idx + 1, len(row_series)):
            val = row_series.iloc[c]
            if pd.notna(val) and str(val).strip() != "":
                match = re.search(r"(\d{4})[-/](\d{2})", str(val))
                if match:
                    anio_mes = f"{match.group(1)}-{match.group(2)}"
                break
        break

data_extraida = []

# 3. Buscar la fila de los días de manera dinámica
row_dias = None
for r_idx in range(len(df_raw)):
    row_vals = df_raw.iloc[r_idx].values
    non_nulls = [x for x in row_vals if pd.notna(x)]
    if len(non_nulls) >= 15:
        try:
            ints = [int(float(x)) for x in non_nulls[:5]]
            if ints == [1, 2, 3, 4, 5]:
                row_dias = r_idx
                break
        except (ValueError, TypeError):
            continue

# 4. Mapear las columnas que corresponden a los días
day_columns = {}
if row_dias is not None:
    for col_idx in range(df_raw.shape[1]):
        val = df_raw.iloc[row_dias, col_idx]
        if pd.notna(val):
            try:
                day_num = int(float(val))
                if 1 <= day_num <= 31:
                    day_columns[col_idx] = day_num
            except (ValueError, TypeError):
                pass

# 5. Extraer uno por uno los datos de cada personal
# Primero buscamos en qué fila empiezan los registros
start_idx = 0
for r in range(len(df_raw)):
    if df_raw.iloc[r].astype(str).str.contains("ID:", na=False).any():
        start_idx = r
        break


# El programa busca hasta que encuentre filas vacías (donde ya no hay "ID:")
row_idx = start_idx
while row_idx < len(df_raw):
    row_series = df_raw.iloc[row_idx]
    
    id_labels = row_series[row_series.astype(str).str.contains("ID:", na=False)]

    if id_labels.empty:
        break
    
    id_col_idx = id_labels.index[0]
    # Buscar el ID (siguiente valor no vacío en la fila)
    emp_id = None
    for c in range(id_col_idx + 1, len(row_series)):
        val = row_series.iloc[c]
        if pd.notna(val) and str(val).strip() != "":
            emp_id = val
            break
    
    # Buscar el Nombre (siguiente valor no vacío tras "Nombre:")
    nombre_labels = row_series[row_series.astype(str).str.contains("Nombre:", na=False)]
    emp_nombre = None
    if not nombre_labels.empty:
        nombre_col_idx = nombre_labels.index[0]
        for c in range(nombre_col_idx + 1, len(row_series)):
            val = row_series.iloc[c]
            if pd.notna(val) and str(val).strip() != "":
                emp_nombre = val
                break
    
    # Las marcaciones están en la fila de abajo
    row_marcaciones = row_idx + 1
    if row_marcaciones < len(df_raw):
        for col_idx, dia in day_columns.items():
            celda_hora = df_raw.iloc[row_marcaciones, col_idx]
            
            if pd.notna(celda_hora) and str(celda_hora).strip() != "":
                contenido = str(celda_hora).strip()
                
                # Separamos las horas usando expresiones regulares (formato HH:MM)
                horas = re.findall(r"\d{2}:\d{2}", contenido)
                
                entrada = horas[0] if len(horas) > 0 else None
                salida = horas[1] if len(horas) > 1 else None
                
                # Formato de fecha completo: AAAA-MM-DD
                fecha_completa = f"{anio_mes}-{str(dia).zfill(2)}"
                
                horas_extras, horas_debe = calcular_horas(fecha_completa, entrada, salida)
                   
                data_extraida.append({
                    "Fecha": fecha_completa,
                    "ID": emp_id,
                    "Nombre": emp_nombre,
                    "Entrada_Raw": a_tiempo(entrada),
                    "Salida_Raw": a_tiempo(salida),
                    "Horas_Extras": horas_extras,
                    "Horas_Debe": horas_debe
                })
    
    # Avanzamos al siguiente bloque de personal (una fila de ID y una de marcaciones = 2)
    row_idx += 2

# 6. Reestructurar el DataFrame a formato horizontal
dates = sorted(list(set([d["Fecha"] for d in data_extraida])))
names = []
for d in data_extraida:
    if d["Nombre"] not in names:
        names.append(d["Nombre"])

cols = ["Día", "Fecha", " "]
for name in names:
    cols.extend([name, "Ingreso", "Salida", "H. Extras", "H. Debe", " "])

rows = [cols]
for date in dates:
    date_obj = datetime.datetime.strptime(date, "%Y-%m-%d")
    days_es = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
    dia_str = days_es[date_obj.weekday()]
    
    row = [dia_str, date, ""]
    for name in names:
        entry = next((item for item in data_extraida if item["Fecha"] == date and item["Nombre"] == name), None)
        if entry:
            row.extend([
                name,
                entry.get("Entrada_Raw", ""),
                entry.get("Salida_Raw", ""),
                entry.get("Horas_Extras", ""),
                entry.get("Horas_Debe", ""),
                "" 
            ])
        else:
            row.extend([name, "", "", "", "", ""])
    rows.append(row)

# Generar fila de Totales
from openpyxl.utils import get_column_letter

total_row = ["Totales", "", ""]
last_data_row_excel = len(dates) + 1

for i, name in enumerate(names):
    extras_idx_excel = 7 + i*6
    debe_idx_excel = 8 + i*6
    
    letra_extras = get_column_letter(extras_idx_excel)
    letra_debe = get_column_letter(debe_idx_excel)
    
    formula_extras = f"=SUM({letra_extras}2:{letra_extras}{last_data_row_excel})"
    formula_debe = f"=SUM({letra_debe}2:{letra_debe}{last_data_row_excel})"
    
    total_row.extend(["", "", "", formula_extras, formula_debe, ""])

rows.append(total_row)
df_wide = pd.DataFrame(rows)

vista_previa_y_guardar(df_wide)
root.mainloop()
